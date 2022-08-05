import openpyxl

# File-->Workbook-->Rows-->Cells

file = "C:\\Users\\lukesh\\PycharmProjects\\PythonTraining\\chapter_14\\data.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row  # count number of rows
cols = sheet.max_column  # count number of columns in Excel sheet

# Reading all the rows & columns from Excel sheet
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(r, c).value, end='            ')
    print()
