import openpyxl

# # same data
# file = "C:\\Users\\lukesh\\PycharmProjects\\PythonTraining\\chapter_14\\testdata.xlsx"
# workbook = openpyxl.load_workbook(file)
# sheet = workbook.active  # (or) sheet = workbook["Data"]   ----- get active sheet from excel
# for r in range(1, 6):
#     for c in range(1, 4):
#         sheet.cell(r, c).value = "welcome"
#
# workbook.save(file)

# multiple data

file = "C:\\Users\\lukesh\\PycharmProjects\\PythonTraining\\chapter_14\\testdata.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active  # (or) sheet = workbook["Data"]   ----- get active sheet from excel

sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "lukesh"
sheet.cell(1, 3).value = "engineer"

sheet.cell(2, 1).value = 222
sheet.cell(2, 2).value = "anuja"
sheet.cell(2, 3).value = "test engineer"

sheet.cell(3, 1).value = "555"
sheet.cell(3, 2).value = "anjali"
sheet.cell(3, 3).value = "subject matter expert"

workbook.save(file)  # save the file after entering the data

